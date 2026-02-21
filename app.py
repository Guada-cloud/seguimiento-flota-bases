# app.py — Pegado directo de Planificación y Realidad · Detección robusta · Español · Export Excel con gráficos
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import re
from io import StringIO, BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart

# ==========================
# Apariencia
# ==========================
st.set_page_config(page_title="Plan vs Real — Operación (Pegado directo)", layout="wide")
TEMPLATE = "plotly_dark"
FONT = "Inter, system-ui, Segoe UI, Roboto"

def stylize(fig, title=None, y_pct=False):
    fig.update_layout(
        template=TEMPLATE, title=title,
        font=dict(family=FONT, size=13, color="#E5E7EB"),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        legend_title_text="", margin=dict(t=45, r=10, b=30, l=10),
    )
    if y_pct: fig.update_yaxes(tickformat=".0%")
    fig.update_xaxes(showgrid=False)
    fig.update_yaxes(gridcolor="rgba(148,163,184,.25)")
    return fig

# ==========================
# Persistencia
# ==========================
DATA_DIR = Path("data"); DATA_DIR.mkdir(exist_ok=True)
BASES_DIR = DATA_DIR / "bases"; BASES_DIR.mkdir(exist_ok=True)
MERGED_CSV = DATA_DIR / "merged.csv"

def save_csv(df: pd.DataFrame, path: Path):
    df.to_csv(path, index=False, encoding="utf-8")

def load_csv(path: Path) -> pd.DataFrame|None:
    return pd.read_csv(path, encoding="utf-8") if path.exists() else None

# ==========================
# Estado
# ==========================
if "bases" not in st.session_state:    # { base: df (varias fechas) }
    st.session_state["bases"] = {}
if "merged" not in st.session_state:
    st.session_state["merged"] = pd.DataFrame()
if "_preview" not in st.session_state: # última previsualización fusionada (plan+real) para base/fecha
    st.session_state["_preview"] = pd.DataFrame()

# ============================================================
# Detección robusta de columnas (muchas variaciones en español)
# ============================================================
SYN = {
    "hora": [
        "hora","hr","tiempo","h"
    ],
    "svc_plan": [
        "svc proy","servicios proy","servicios proyectados","svc plan","serv plan",
        "proyectado","planificado","plan","proy","proyectados"
    ],
    "svc_real": [
        "svc reales","servicios reales","svc real","serv real","real","reales",
        "observado","observados"
    ],
    "mov_plan": [
        # requeridos / planificados
        "mov req","mov requeridos","mov plan","moviles plan","móviles plan","moviles requeridos",
        "req moviles","requeridos","plan moviles","movil requerido","moviles requeridas",
        "dotación plan","dotacion plan","staff plan","agentes plan","operadores plan"
    ],
    "mov_real": [
        # nómina / reales / dotación efectiva
        "moviles x nomina","mov x nomina","moviles nomina","movil nomina","nómina","nomina",
        "mov reales","mov real","móviles reales","móviles real",
        "dotación","dotacion","staff","agentes","operadores","plantilla","planta",
        "dotación efectiva","dotacion efectiva"
    ],
    "coef_hs": [
        "coeficiente hs","coef hs","coef hs.","coeficiente horas","coeficiente segun hs op previstas"
    ],
    "dif_mov": [
        "dif moviles","dif mov","delta moviles","delta mov.","variacion moviles","variación móviles"
    ]
}

# si no viene encabezado, asigno por posición típica
DEFAULT_ORDER = ["hora","svc_plan","svc_real","mov_plan","mov_real","coef_hs","dif_mov"]

def _norm(s: str) -> str:
    s = str(s).strip().lower()
    rep = {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n"}
    for a,b in rep.items(): s = s.replace(a,b)
    s = re.sub(r"\s+"," ", s)
    return s

def _find_col(cols, aliases):
    m = { _norm(c): c for c in cols }
    # exacto
    for a in aliases:
        if a in m: return m[a]
    # contiene
    for a in aliases:
        for k,v in m.items():
            if a in k: return v
    return None

def _smart_sep(text: str) -> str:
    sample = text[:1000]
    if "\t" in sample: return "\t"
    if ";" in sample:  return ";"
    return ","  # fallback

def _to_num_series(s: pd.Series) -> pd.Series:
    s = s.astype(str).str.strip()
    s = s.replace({"": np.nan, "nan": np.nan, "None": np.nan,
                   "#¿NOMBRE?": np.nan, "#¡NOMBRE?": np.nan, "#VALUE!": np.nan}, regex=False)
    def _fix_one(x:str):
        if x is np.nan or x is None: return np.nan
        txt = str(x)
        # miles/decimal heurística
        if "," in txt and "." in txt:
            # última coma > último punto => coma decimal
            if txt.rfind(",") > txt.rfind("."):
                txt = txt.replace(".", "").replace(",", ".")
            else:
                txt = txt.replace(",", "")
        elif "," in txt:
            txt = txt.replace(",", ".")
        try:
            return float(txt)
        except:
            return np.nan
    return s.map(_fix_one)

def _to_time_series(s: pd.Series) -> pd.Series:
    # Acepta "HH:MM", "H:MM" o serial Excel (si viniesen números)
    try:
        return pd.to_datetime(s.astype(str), errors="coerce").dt.time
    except:
        return pd.to_datetime([""], errors="coerce").dt.time  # todo NaT

def parse_pasted_flexible(text: str) -> pd.DataFrame:
    """
    Parser flexible: acepta encabezados "casi cualquiera" o sin encabezados.
    Devuelve columnas estandar: Hora, SvcPlan, SvcReal, MovPlan, MovReal, CoefHS, DifMov_archivo
    """
    if not text or not text.strip():
        return pd.DataFrame()
    sep = _smart_sep(text)

    # intento con encabezado
    df = pd.read_csv(StringIO(text), sep=sep, engine="python", dtype=str)
    cols = list(df.columns)

    # ¿reconocemos HORA?
    hora_c = _find_col(cols, SYN["hora"])

    if hora_c is None:
        # puede venir SIN encabezado → re-leo sin header
        df = pd.read_csv(StringIO(text), sep=sep, header=None, engine="python", dtype=str)
        # ubico la col de hora por patrón HH:MM en la primera fila
        guess_hora_idx = None
        for j in range(min(10, df.shape[1])): # miro primeras 10
            val = str(df.iloc[0, j])
            if re.match(r"^\d{1,2}:\d{2}$", val.strip()):
                guess_hora_idx = j; break
        if guess_hora_idx is None:
            # como último recurso, pruebo convertir todas y ver cuál tiene menos NaT
            best, best_j = 9999, None
            for j in range(df.shape[1]):
                ok = pd.to_datetime(df[j].astype(str), errors="coerce").dt.time.notna().sum()
                if ok < best:
                    best = ok; best_j = j
            guess_hora_idx = best_j

        # asigno encabezados por orden típico
        n = df.shape[1]
        names = DEFAULT_ORDER[:n]
        df.columns = names
        # renombro la col detectada como 'hora'
        if guess_hora_idx is not None and guess_hora_idx < len(names):
            names[guess_hora_idx] = "hora"
            df.columns = names

        # mapeo a estándar
        out = pd.DataFrame()
        out["Hora"]    = _to_time_series(df.get("hora", pd.Series(dtype=str)))
        out["SvcPlan"] = _to_num_series(df.get("svc_plan", pd.Series(dtype=str)))
        out["SvcReal"] = _to_num_series(df.get("svc_real", pd.Series(dtype=str)))
        out["MovPlan"] = _to_num_series(df.get("mov_plan", pd.Series(dtype=str)))
        out["MovReal"] = _to_num_series(df.get("mov_real", pd.Series(dtype=str)))
        out["CoefHS"]  = _to_num_series(df.get("coef_hs", pd.Series(dtype=str)))
        out["DifMov_archivo"] = _to_num_series(df.get("dif_mov", pd.Series(dtype=str)))
        out = out[out["Hora"].notna()]
        out["HoraStr"] = pd.to_datetime(out["Hora"].astype(str)).dt.strftime("%H:%M")
        return out.reset_index(drop=True)

    # con encabezado → mapeo flexible
    sp_c   = _find_col(cols, SYN["svc_plan"])
    sr_c   = _find_col(cols, SYN["svc_real"])
    mp_c   = _find_col(cols, SYN["mov_plan"])
    mr_c   = _find_col(cols, SYN["mov_real"])
    coef_c = _find_col(cols, SYN["coef_hs"])
    dif_c  = _find_col(cols, SYN["dif_mov"])

    out = pd.DataFrame()
    out["Hora"]    = _to_time_series(df[hora_c])
    out["SvcPlan"] = _to_num_series(df[sp_c]) if sp_c else np.nan
    out["SvcReal"] = _to_num_series(df[sr_c]) if sr_c else np.nan
    out["MovPlan"] = _to_num_series(df[mp_c]) if mp_c else np.nan
    out["MovReal"] = _to_num_series(df[mr_c]) if mr_c else np.nan
    out["CoefHS"]  = _to_num_series(df[coef_c]) if coef_c else np.nan
    out["DifMov_archivo"] = _to_num_series(df[dif_c]) if dif_c else np.nan
    out = out[out["Hora"].notna()]
    out["HoraStr"] = pd.to_datetime(out["Hora"].astype(str)).dt.strftime("%H:%M")
    return out.reset_index(drop=True)

def merge_plan_real(plan_df: pd.DataFrame, real_df: pd.DataFrame) -> pd.DataFrame:
    """
    Fusiona por Hora. Si alguno trae columnas del otro (p.ej. pegaste todo en uno), se respeta.
    """
    # normalizo claves mínimas
    left  = plan_df[["Hora","HoraStr"]].copy()
    for c in ["SvcPlan","MovPlan","CoefHS","DifMov_archivo"]:
        left[c] = plan_df[c] if c in plan_df else np.nan

    right = real_df[["HoraStr"]].copy()
    for c in ["SvcReal","MovReal"]:
        right[c] = real_df[c] if c in real_df else np.nan

    m = pd.merge(left, right, on="HoraStr", how="outer")
    m["Hora"] = m["Hora"].fillna(pd.to_datetime(m["HoraStr"]).dt.time)

    # si en plan venía algo de real (o viceversa), completo con el no nulo
    if "SvcPlan" in plan_df and "SvcReal" in plan_df:
        m["SvcReal"] = np.where(m["SvcReal"].notna(), m["SvcReal"], plan_df.set_index("HoraStr")["SvcReal"].reindex(m["HoraStr"]).values)
    if "MovPlan" in plan_df and "MovReal" in plan_df:
        m["MovReal"] = np.where(m["MovReal"].notna(), m["MovReal"], plan_df.set_index("HoraStr")["MovReal"].reindex(m["HoraStr"]).values)

    # Dif móviles del archivo o calculado
    dif_calc = m["MovReal"] - m["MovPlan"]
    m["DifMov"] = np.where(m.get("DifMov_archivo", pd.Series([np.nan]*len(m))).notna(), m["DifMov_archivo"], dif_calc)

    out = pd.DataFrame({
        "Hora": m["Hora"], "HoraStr": m["HoraStr"],
        "Servicios_Planificados": m["SvcPlan"],
        "Servicios_Reales": m["SvcReal"],
        "Moviles_Planificados": m["MovPlan"],
        "Moviles_Reales": m["MovReal"],
        "Coeficiente_HS": m["CoefHS"],
        "Dif_Moviles": m["DifMov"]
    }).sort_values("HoraStr").reset_index(drop=True)
    return out

def enrich_with_time_and_metrics(df: pd.DataFrame, fecha, base) -> pd.DataFrame:
    out = df.copy()
    out["Fecha"] = pd.to_datetime(str(fecha)).date()
    out["Base"]  = str(base).strip().upper()

    out["Fecha_dt"] = pd.to_datetime(out["Fecha"])
    iso = out["Fecha_dt"].dt.isocalendar()
    out["Año"]    = out["Fecha_dt"].dt.year
    out["Mes"]    = out["Fecha_dt"].dt.month
    out["Semana"] = iso.week
    out["Dia"]    = out["Fecha_dt"].dt.day

    # Métricas
    out["Dif_Servicios"] = out["Servicios_Reales"] - out["Servicios_Planificados"]
    out["Desvio_Servicios_%"] = np.where(out["Servicios_Planificados"]>0, out["Dif_Servicios"]/out["Servicios_Planificados"]*100, np.nan)
    out["Desvio_Moviles_%"]   = np.where(out["Moviles_Planificados"]>0, out["Dif_Moviles"]/out["Moviles_Planificados"]*100, np.nan)
    out["Efectividad"] = np.where(out["Servicios_Planificados"]>0,
                                  1 - (out["Dif_Servicios"].abs()/out["Servicios_Planificados"]), np.nan)
    out["APE"] = np.where(out["Servicios_Planificados"]>0,
                          (out["Servicios_Reales"] - out["Servicios_Planificados"]).abs()/out["Servicios_Planificados"], np.nan)
    out["AE"]  = (out["Servicios_Reales"] - out["Servicios_Planificados"]).abs()
    out["Bias"]= (out["Servicios_Planificados"] - out["Servicios_Reales"])

    out["Status"] = np.select(
        [out["Servicios_Planificados"].notna() & out["Servicios_Reales"].isna(),
         out["Servicios_Planificados"].isna() & out["Servicios_Reales"].notna()],
        ["No ejecutado","No planificado"], default="OK"
    )
    out["Clasificacion"] = np.select(
        [out["Status"].eq("No ejecutado"),
         out["Status"].eq("No planificado"),
         out["Dif_Servicios"].fillna(0).eq(0),
         out["Dif_Servicios"].fillna(0)>0,
         out["Dif_Servicios"].fillna(0)<0],
        ["No ejecutado","No planificado","Exacto","Sobre planificado","Bajo planificado"], default="NA"
    )
    return out

# ==========================
# Export Excel enriquecido (bonito, con gráficos y explicación)
# ==========================
def export_excel_pretty(df: pd.DataFrame, nombre="analisis_plan_vs_real.xlsx"):
    # Hojas: Resumen (KPIs + texto), Graficos (2 charts), Datos (tabla)
    wb = Workbook()
    ws_res = wb.active; ws_res.title = "Resumen"
    ws_g   = wb.create_sheet("Gráficos")
    ws_d   = wb.create_sheet("Datos")

    # --------- Estilos ----------
    thin = Side(style="thin", color="334155")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_hdr = PatternFill("solid", fgColor="0B1221")
    fill_blk = PatternFill("solid", fgColor="111827")
    font_hdr = Font(color="E5E7EB", bold=True, size=13)
    font_txt = Font(color="E5E7EB")
    # Fondo oscuro
    for ws in [ws_res, ws_g, ws_d]:
        for col in range(1, 12):
            ws.column_dimensions[chr(64+col if col<=26 else 64)].width = 16

    # --------- Resumen ----------
    k1 = ("Efectividad de la planificación",)
    k2 = ("% Desvío en Servicios",)
    k3 = ("% Desvío en Móviles",)

    tot_plan_s = df["Servicios_Planificados"].sum()
    tot_real_s = df["Servicios_Reales"].sum()
    tot_plan_m = df["Moviles_Planificados"].sum()
    tot_real_m = df["Moviles_Reales"].sum()

    efect = 1 - (abs(tot_real_s - tot_plan_s)/tot_plan_s) if tot_plan_s>0 else np.nan
    desv_s = (tot_real_s - tot_plan_s)/tot_plan_s*100 if tot_plan_s>0 else np.nan
    desv_m = (tot_real_m - tot_plan_m)/tot_plan_m*100 if tot_plan_m>0 else np.nan

    rows = [
        ["Indicador", "Valor"],
        ["Efectividad de la planificación", f"{efect:.2%}" if pd.notna(efect) else "—"],
        ["% Desvío Servicios (Real - Plan)", f"{desv_s:,.1f}%" if pd.notna(desv_s) else "—"],
        ["% Desvío Móviles (Real - Plan)", f"{desv_m:,.1f}%" if pd.notna(desv_m) else "—"],
        ["MAPE (Servicios)", f"{(df['APE'].mean()*100):.1f}%" if df['APE'].notna().any() else "—"],
        ["MAE (Servicios)", f"{df['AE'].mean():.2f}" if df['AE'].notna().any() else "—"],
        ["Forecast Bias (Servicios)", f"{(df['Bias'].sum()/df['Servicios_Reales'].sum()*100):.1f}%" if df['Servicios_Reales'].sum()!=0 else "—"],
    ]
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            c = ws_res.cell(row=i, column=j, value=val)
            c.font = font_hdr if i==1 or j==1 else font_txt
            c.fill = fill_hdr if i==1 or j==1 else fill_blk
            c.alignment = Alignment(horizontal="left")
            c.border = border

    # Explicación
    exp_txt = (
        "Efectividad = 1 − |Real − Plan| / Plan (sobre Servicios).  "
        "Desvío % = (Real − Plan) / Plan × 100.  "
        "MAPE/MAE/Bias calculados sobre Servicios.  "
        "Los valores están filtrados según la selección de la app."
    )
    ws_res.cell(row=len(rows)+2, column=1, value=exp_txt).font = font_txt

    # --------- Datos ----------
    # Escribo el df en la hoja "Datos"
    cols = list(df.columns)
    for j, c in enumerate(cols, start=1):
        cell = ws_d.cell(row=1, column=j, value=c)
        cell.font = font_hdr; cell.fill = fill_hdr; cell.border = border
    for i, r in enumerate(df.itertuples(index=False), start=2):
        for j, v in enumerate(r, start=1):
            cell = ws_d.cell(row=i, column=j, value=v)
            cell.font = font_txt; cell.fill = fill_blk; cell.border = border

    # --------- Gráficos ----------
    # Gráfico 1: Línea Plan vs Real (Servicios)
    # Ubico series en columnas de "Datos"
    def col_idx(name):
        return cols.index(name)+1 if name in cols else None
    x_col = col_idx("HoraStr")
    y1    = col_idx("Servicios_Planificados")
    y2    = col_idx("Servicios_Reales")
    ydv   = col_idx("Dif_Servicios")

    max_row = len(df)+1

    if x_col and y1 and y2:
        lc = LineChart()
        lc.title = "Plan vs Real (Servicios por hora)"
        lc.style = 12
        data1 = Reference(ws_d, min_col=y1, min_row=1, max_row=max_row)  # incluye encabezado
        data2 = Reference(ws_d, min_col=y2, min_row=1, max_row=max_row)
        cats  = Reference(ws_d, min_col=x_col, min_row=2, max_row=max_row)
        lc.add_data(data1, titles_from_data=True)
        lc.add_data(data2, titles_from_data=True)
        lc.set_categories(cats)
        ws_g.add_chart(lc, "A2")

    # Gráfico 2: Barras Desvío % por hora (Servicios)
    if x_col and ydv:
        bc = BarChart()
        bc.title = "Desvío (Servicios) por hora — Real − Plan"
        bc.style = 12
        data3 = Reference(ws_d, min_col=ydv, min_row=1, max_row=max_row)
        cats  = Reference(ws_d, min_col=x_col, min_row=2, max_row=max_row)
        bc.add_data(data3, titles_from_data=True)
        bc.set_categories(cats)
        ws_g.add_chart(bc, "A20")

    # Salida
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue(), nombre

# ==========================
# Sidebar — Pegar Plan y Real
# ==========================
with st.sidebar:
    st.header("Pegar datos — Planificación y Realidad (sin Excel)")
    bases_exist = sorted(st.session_state["bases"].keys())
    base_sel = st.selectbox("Base", options=["(nueva)"] + bases_exist, index=0)
    base_name = st.text_input("Nombre de Base", value="" if base_sel=="(nueva)" else base_sel, help="Ej.: PROY_6001 / DEMOTOS")
    fecha_in = st.date_input("Fecha", value=None)

    st.caption("Pegá tablas copiadas desde Excel. Acepta TAB/;/,; decimales con coma o punto; encabezados variables o sin encabezado.")
    with st.expander("Planificación (pegar tabla que contenga SVC PROY y/o MOV REQ)", expanded=True):
        txt_plan = st.text_area("Pegar Planificación", height=170, key="paste_plan")
    with st.expander("Realidad (pegar tabla que contenga SVC REALES y/o MOVILES X NÓMINA)", expanded=True):
        txt_real = st.text_area("Pegar Realidad", height=170, key="paste_real")

    b1, b2 = st.columns(2)
    with b1:
        if st.button("🔎 Previsualizar y fusionar"):
            if not base_name:
                st.error("Ingresá nombre de Base."); st.stop()
            if not fecha_in:
                st.error("Elegí la Fecha."); st.stop()
            try:
                df_p = parse_pasted_flexible(txt_plan) if txt_plan.strip() else pd.DataFrame()
                df_r = parse_pasted_flexible(txt_real) if txt_real.strip() else pd.DataFrame()
                if df_p.empty and df_r.empty:
                    st.error("No hay datos en Plan ni en Real."); st.stop()
                if df_p.empty: df_p = df_r.copy()
                if df_r.empty: df_r = df_p.copy()

                fused = merge_plan_real(df_p, df_r)
                prev = enrich_with_time_and_metrics(fused, fecha_in, base_name)
                st.session_state["_preview"] = prev
                st.success(f"Previsualización OK — filas: {len(prev)}")
                st.dataframe(prev.head(24), use_container_width=True)
            except Exception as e:
                st.error(f"No se pudo leer/fusionar: {e}")

    with b2:
        if st.button("💾 Guardar Base (día)"):
            if st.session_state["_preview"].empty:
                st.info("Primero presioná 'Previsualizar y fusionar'.")
            else:
                df_prev = st.session_state["bases"].get(base_name, pd.DataFrame())
                if not df_prev.empty:
                    df_prev = df_prev[~df_prev["Fecha"].eq(pd.to_datetime(fecha_in).date())]
                    df_new = pd.concat([df_prev, st.session_state["_preview"]], ignore_index=True)
                else:
                    df_new = st.session_state["_preview"].copy()
                st.session_state["bases"][base_name] = df_new
                save_csv(df_new, BASES_DIR / f"{base_name}.csv")
                st.success(f"Base '{base_name}' guardada ({len(df_new)} filas totales).")

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("💾 Guardar MERGED (/data)"):
            dfs = [df.copy() for df in st.session_state["bases"].values() if not df.empty]
            merged = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
            st.session_state["merged"] = merged
            if not merged.empty:
                save_csv(merged, MERGED_CSV); st.success(f"MERGED guardado: {len(merged):,} filas.")
            else:
                st.info("No hay datos para guardar.")
    with c2:
        if st.button("🧹 Limpiar memoria"):
            st.session_state["bases"] = {}
            st.session_state["merged"] = pd.DataFrame()
            st.session_state["_preview"] = pd.DataFrame()
            st.success("Memoria limpiada (no borra /data).")

# ==========================
# Área principal — Tabs + Filtros
# ==========================
st.title("Análisis de Planificación vs Realidad (pegado directo)")
tabs = st.tabs(["Tablero", "Análisis por Base", "Análisis Horario", "Auditoría Detallada"])

# Dataset unificado (vivo)
dfs_live = [df.copy() for df in st.session_state["bases"].values() if not df.empty]
merged_live = pd.concat(dfs_live, ignore_index=True) if dfs_live else pd.DataFrame()
st.session_state["merged"] = merged_live

# Filtros
fl = st.container()
with fl:
    c1,c2,c3,c4,c5,c6 = st.columns([1.4,1,1,1,1.2,1.6])
    with c1:
        bases_all = sorted(merged_live["Base"].unique().tolist()) if not merged_live.empty else []
        bases_fil = st.multiselect("Base", options=bases_all, default=bases_all)
    with c2:
        fecha_fil = st.date_input("Día", value=None, key="dia_filter")
    with c3:
        semana_fil = st.number_input("Semana (ISO)", value=0, step=1, min_value=0)
    with c4:
        anio_fil = st.number_input("Año", value=0, step=1, min_value=0)
    with c5:
        mes_fil = st.number_input("Mes", value=0, step=1, min_value=0, help="1-12 (opcional).")
    with c6:
        horas_all = sorted(merged_live["HoraStr"].unique().tolist()) if not merged_live.empty else []
        horas_fil = st.multiselect("Hora (HH:MM)", options=horas_all, default=horas_all)

def apply_filters(df):
    d = df.copy()
    if bases_fil: d = d[d["Base"].isin(bases_fil)]
    if fecha_fil is not None: d = d[d["Fecha"].eq(pd.to_datetime(fecha_fil).date())]
    if semana_fil and semana_fil>0: d = d[d["Semana"].eq(int(semana_fil))]
    if anio_fil and anio_fil>0: d = d[d["Año"].eq(int(anio_fil))]
    if mes_fil and mes_fil>0: d = d[d["Mes"].eq(int(mes_fil))]
    if horas_fil: d = d[d["HoraStr"].isin(horas_fil)]
    return d

# ==========================
# TAB 1 — Tablero
# ==========================
with tabs[0]:
    df_f = apply_filters(merged_live)
    if df_f.empty:
        st.info("Pegá/guardá datos en el lateral y/o ajustá filtros.")
    else:
        st.subheader("Indicadores generales")
        tot_plan_s = df_f["Servicios_Planificados"].sum()
        tot_real_s = df_f["Servicios_Reales"].sum()
        tot_plan_m = df_f["Moviles_Planificados"].sum()
        tot_real_m = df_f["Moviles_Reales"].sum()

        desvio_s = (tot_real_s - tot_plan_s)/tot_plan_s*100 if tot_plan_s>0 else np.nan
        desvio_m = (tot_real_m - tot_plan_m)/tot_plan_m*100 if tot_plan_m>0 else np.nan
        efect    = 1 - (abs(tot_real_s - tot_plan_s)/tot_plan_s) if tot_plan_s>0 else np.nan

        k1,k2,k3,k4 = st.columns(4)
        k1.metric("Efectividad de la planificación", f"{efect:.1%}" if pd.notna(efect) else "—")
        k2.metric("Desvío Servicios (%)", f"{desvio_s:,.1f}%" if pd.notna(desvio_s) else "—")
        k3.metric("Desvío Móviles (%)", f"{desvio_m:,.1f}%" if pd.notna(desvio_m) else "—")
        k4.metric("Sesgo (Bias) Servicios (%)",
                  f"{(df_f['Bias'].sum()/df_f['Servicios_Reales'].sum()*100):.1f}%" if df_f["Servicios_Reales"].sum()!=0 else "—")

        # Gráfico 1: Diferencia de Servicios por hora
        g1 = df_f.groupby("HoraStr", as_index=False)[["Servicios_Planificados","Servicios_Reales","Dif_Servicios"]].sum()
        fig1 = px.bar(g1, x="HoraStr", y="Dif_Servicios", color="Dif_Servicios",
                      color_continuous_scale="RdYlGn", title="Desvío de Servicios por hora (Real − Plan)")
        stylize(fig1); st.plotly_chart(fig1, use_container_width=True)

        # Gráfico 2: Diferencia de Móviles por hora
        g2 = df_f.groupby("HoraStr", as_index=False)[["Moviles_Planificados","Moviles_Reales","Dif_Moviles"]].sum()
        fig2 = px.bar(g2, x="HoraStr", y="Dif_Moviles", color="Dif_Moviles",
                      color_continuous_scale="RdYlGn", title="Desvío de Móviles por hora (Real − Plan)")
        stylize(fig2); st.plotly_chart(fig2, use_container_width=True)

        # Mapa de calor (Servicios)
        piv = df_f.pivot_table(values="Dif_Servicios", index="Fecha", columns="HoraStr", aggfunc="sum").fillna(0)
        if not piv.empty:
            fig3 = px.imshow(piv, color_continuous_scale="RdYlGn", aspect="auto",
                             title="Mapa de calor — Desvío de servicios (Real − Plan)")
            stylize(fig3); st.plotly_chart(fig3, use_container_width=True)

        # Top‑5 y Base con mayor desvío
        hsum = df_f.groupby("HoraStr", as_index=False)["Dif_Servicios"].sum()
        sub  = hsum.nsmallest(5, "Dif_Servicios")
        sobre= hsum.nlargest(5, "Dif_Servicios")
        wb   = df_f.groupby("Base", as_index=False)["Dif_Servicios"].apply(lambda s: s.abs().sum()) \
                   .rename(columns={"Dif_Servicios":"Desvío absoluto"}).sort_values("Desvío absoluto", ascending=False).head(1)
        c1,c2,c3 = st.columns(3)
        with c1: st.subheader("Top 5 Sub‑plan (Servicios)"); st.dataframe(sub, use_container_width=True, hide_index=True)
        with c2: st.subheader("Top 5 Sobre‑plan (Servicios)"); st.dataframe(sobre, use_container_width=True, hide_index=True)
        with c3: st.subheader("Base con mayor desvío"); st.dataframe(wb, use_container_width=True, hide_index=True)

# ==========================
# TAB 2 — Análisis por Base
# ==========================
with tabs[1]:
    df_f = apply_filters(merged_live)
    if df_f.empty:
        st.info("No hay datos para los filtros seleccionados.")
    else:
        st.subheader("Desvío por Base (Servicios)")
        g = df_f.groupby("Base", as_index=False)[["Servicios_Planificados","Servicios_Reales"]].sum()
        g["Desvío_%"] = np.where(g["Servicios_Planificados"]>0,
                                 (g["Servicios_Reales"]-g["Servicios_Planificados"])/g["Servicios_Planificados"]*100, np.nan)
        fig = px.bar(g, x="Base", y="Desvío_%", color="Desvío_%", color_continuous_scale="RdYlGn",
                     title="Desvío % por Base (Servicios)")
        stylize(fig); st.plotly_chart(fig, use_container_width=True)
        st.dataframe(g, use_container_width=True, hide_index=True)

# ==========================
# TAB 3 — Análisis Horario
# ==========================
with tabs[2]:
    df_f = apply_filters(merged_live)
    if df_f.empty:
        st.info("No hay datos para los filtros seleccionados.")
    else:
        st.subheader("Series por hora — Plan vs Real")
        g = df_f.groupby("HoraStr", as_index=False)[["Servicios_Planificados","Servicios_Reales",
                                                     "Moviles_Planificados","Moviles_Reales"]].sum()
        fig = px.line(g, x="HoraStr", y=["Servicios_Planificados","Servicios_Reales"], title="Servicios — Plan vs Real")
        stylize(fig); st.plotly_chart(fig, use_container_width=True)
        figm = px.line(g, x="HoraStr", y=["Moviles_Planificados","Moviles_Reales"], title="Móviles — Plan vs Real")
        stylize(figm); st.plotly_chart(figm, use_container_width=True)

        st.subheader("Tabla horaria (detalle)")
        st.dataframe(
            df_f[["Fecha","HoraStr","Base","Servicios_Planificados","Servicios_Reales","Dif_Servicios",
                  "Moviles_Planificados","Moviles_Reales","Dif_Moviles","Efectividad","Clasificacion"]]
            .sort_values(["Fecha","HoraStr","Base"]),
            use_container_width=True, hide_index=True
        )

# ==========================
# TAB 4 — Auditoría Detallada (descarga Excel lindo)
# ==========================
with tabs[3]:
    df_f = apply_filters(merged_live)
    if df_f.empty:
        st.info("No hay datos para los filtros seleccionados.")
    else:
        st.subheader("Auditoría (lo que estás viendo)")
        cols = ["Fecha","HoraStr","Base",
                "Servicios_Planificados","Servicios_Reales","Dif_Servicios","Desvio_Servicios_%",
                "Moviles_Planificados","Moviles_Reales","Dif_Moviles","Desvio_Moviles_%",
                "Efectividad","Clasificacion","Status","Semana","Mes","Año","Coeficiente_HS"]
        cols = [c for c in cols if c in df_f.columns]
        df_aud = df_f[cols].sort_values(["Fecha","HoraStr","Base"])

        st.dataframe(df_aud, use_container_width=True, hide_index=True)
        # Excel enriquecido (KPIs + 2 gráficos + datos)
        xls, fname = export_excel_pretty(df_aud, "analisis_plan_vs_real.xlsx")
        st.download_button("⬇️ Descargar Excel enriquecido", data=xls, file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
